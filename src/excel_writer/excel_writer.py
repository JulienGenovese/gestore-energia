import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    def __init__(
        self,
        df: pd.DataFrame,
        output_path: str,
        key_columns: list[str],
        price_columns: list[str],
        note_column: str | None = None,
    ):
        self.df = df
        self.output_path = output_path

        self.key_columns = list(key_columns)
        self.price_columns = list(price_columns)
        self.note_column = note_column

        self.fills = {
            "key": PatternFill("solid", fgColor="92D050"),
            "price": PatternFill("solid", fgColor="FF0000"),
            "note": PatternFill("solid", fgColor="FFFF00"),
            "other": PatternFill("solid", fgColor="D9D9D9"),
        }

        self.font = Font(name="Calibri")
        self.header_font = Font(name="Calibri", bold=True)

        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_center = Alignment(horizontal="center", vertical="center")

    def write_excel(self):
        self.df.to_excel(self.output_path, index=False)

    def format_sheet(self):
        wb = load_workbook(self.output_path)
        ws = wb.active

        for col_idx, col in enumerate(ws.iter_cols(), start=1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            header = col[0].value

            # Decide colore header
            if header in self.key_columns:
                header_fill = self.fills["key"]
            elif header in self.price_columns:
                header_fill = self.fills["price"]
            elif header == self.note_column:
                header_fill = self.fills["note"]
            else:
                header_fill = self.fills["other"]

            for row_idx, cell in enumerate(col, start=1):
                cell.font = self.header_font if row_idx == 1 else self.font

                if row_idx == 1:
                    cell.fill = header_fill
                    cell.alignment = self.align_center
                else:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = self.align_center
                    else:
                        cell.alignment = self.align_left

                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(self.output_path)

    def run(self):
        self.write_excel()
        self.format_sheet()
